import os
import io
import logging
import email
from email import policy
import pandas as pd
import gspread
from imapclient import IMAPClient
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook


# --- 1. НАСТРОЙКА ЛОГИРОВАНИЯ ---
# Настроим логирование, чтобы видеть, что происходит на Render.
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# --- 2. ПОЛУЧЕНИЕ ПЕРЕМЕННЫХ ОКРУЖЕНИЯ ---
# Все важные данные мы будем хранить в переменных окружения на Render.
# Для локальной проверки их можно добавить в файл .env.
YANDEX_LOGIN = os.getenv("YANDEX_LOGIN")
YANDEX_APP_PASSWORD = os.getenv("YANDEX_APP_PASSWORD")
SENDER_TO_CHECK = os.getenv("SENDER_TO_CHECK")

# Для Google Таблиц
SERVICE_ACCOUNT_JSON = os.getenv("SERVICE_ACCOUNT_JSON") # Это будет строка с JSON-ключом
SPREADSHEET_NAME = os.getenv("SPREADSHEET_NAME")
WORKSHEET_NAME = os.getenv("WORKSHEET_NAME", "Лист1") # По умолчанию "Лист1"

# Проверяем, что все переменные установлены
required_vars = [
    YANDEX_LOGIN, YANDEX_APP_PASSWORD, SENDER_TO_CHECK,
    SERVICE_ACCOUNT_JSON, SPREADSHEET_NAME
]
if not all(required_vars):
    logging.error("Ошибка: Не все переменные окружения установлены!")
    exit(1)

def parse_quotation_excel(file_content: bytes) -> list:
    """
    Парсит содержимое Excel-файла, извлекая данные из таблицы.
    Возвращает список словарей, где каждый словарь — одна строка для Google Sheets.
    """
    try:
        # Открываем файл из байтового потока
        excel_file = io.BytesIO(file_content)
        wb = load_workbook(excel_file, data_only=True)
        
        # Предполагаем, что нужный лист называется "Quotation"
        if "Quotation" not in wb.sheetnames:
            logging.error(f"Лист 'Quotation' не найден. Доступные листы: {wb.sheetnames}")
            return []
        ws = wb["Quotation"]

        # --- Шаг 1: Извлекаем RFQ и Date из шапки ---
        # По вашей структуре, RFQ в H2, Date в O2
        rfq = ws["H2"].value if ws["H2"].value else ""
        date_cell = ws["O2"].value
        if hasattr(date_cell, 'strftime'):
            date_str = date_cell.strftime("%d.%m.%Y")
        else:
            date_str = str(date_cell) if date_cell else ""

        logging.info(f"Извлечены данные шапки: RFQ='{rfq}', Date='{date_str}'")

        # --- Шаг 2: Находим строку с заголовками таблицы ---
        # Ищем строку, где в первом столбце значение "№"
        header_row_idx = None
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == "№":
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            logging.error("Не найдена строка заголовков с '№'")
            return []

        # --- Шаг 3: Читаем строки данных ---
        rows_data = []
        current_row = header_row_idx + 1
        while current_row <= ws.max_row:
            # Проверяем, что в первой ячейке номер позиции (1.0, 2.0...)
            first_val = ws.cell(row=current_row, column=1).value
            if first_val is None or str(first_val).strip() == "":
                break

            # Извлекаем данные по нужным колонкам
            # Используем индексы колонок, как в вашем файле:
            # A:№, B:Part Number, D:Description, E:Alternate PN,
            # F:Qty, G:MU, H:Req.Condition, I:Target Date, T:Remarks (как Comment)
            
            # Обработка числовых значений для Qty
            qty_val = ws.cell(row=current_row, column=6).value # колонка F
            if isinstance(qty_val, (int, float)):
                qty_str = str(int(qty_val)) if qty_val == int(qty_val) else str(qty_val)
            else:
                qty_str = str(qty_val) if qty_val is not None else ""
            
            # Обработка даты
            target_date_val = ws.cell(row=current_row, column=9).value # колонка I
            if hasattr(target_date_val, 'strftime'):
                target_date_str = target_date_val.strftime("%d.%m.%Y")
            else:
                target_date_str = str(target_date_val) if target_date_val else ""
                
            row_dict = {
                "RFQ": rfq,
                "Date": date_str,
                "№": str(first_val).split('.')[0],  # Превращаем "1.0" в "1"
                "PN": ws.cell(row=current_row, column=2).value or "",
                "DESC": ws.cell(row=current_row, column=4).value or "",
                "Alt": ws.cell(row=current_row, column=5).value or "no",
                "R. Qty.": qty_str,
                "Unit": ws.cell(row=current_row, column=7).value or "",
                "Req.Condition": ws.cell(row=current_row, column=8).value or "",
                "Target Date": target_date_str,
                "Comment": ws.cell(row=current_row, column=20).value or ""  # колонка T
            }
            rows_data.append(row_dict)
            current_row += 1

        logging.info(f"Успешно распарсено {len(rows_data)} позиций")
        return rows_data

    except Exception as e:
        logging.error(f"Ошибка при парсинге Excel: {e}", exc_info=True)
        return []

def main():
    logging.info("Начало выполнения скрипта")
    
    # --- Шаг 1: Подключение к Яндекс.Почте и скачивание вложения ---
    try:
        with IMAPClient("imap.yandex.ru") as client:
            logging.info(f"Подключение к почте {YANDEX_LOGIN}...")
            client.login(YANDEX_LOGIN, YANDEX_APP_PASSWORD)
            client.select_folder("INBOX")
            
            # Ищем непрочитанные письма от нужного отправителя
            messages = client.search(['UNSEEN', 'FROM', SENDER_TO_CHECK])
            logging.info(f"Найдено непрочитанных писем от {SENDER_TO_CHECK}: {len(messages)}")
            
            if not messages:
                logging.info("Нет новых писем для обработки.")
                return

            # Обрабатываем каждое новое письмо
            for msg_id in messages:
                logging.info(f"Обработка письма ID: {msg_id}")
                email_data = client.fetch([msg_id], ['BODY.PEEK[]'])
                msg = email.message_from_bytes(email_data[msg_id][b'BODY.PEEK[]'], policy=policy.default)

                # Ищем вложение с расширением .xlsx
                xlsx_content = None
                for part in msg.iter_attachments():
                    filename = part.get_filename()
                    if filename and filename.lower().endswith('.xlsx'):
                        logging.info(f"Найдено вложение: {filename}")
                        xlsx_content = part.get_content()
                        break

                if xlsx_content is None:
                    logging.warning(f"В письме {msg_id} не найдено вложения .xlsx")
                    continue

                # --- Шаг 2: Парсинг вложения ---
                parsed_data = parse_quotation_excel(xlsx_content)
                if not parsed_data:
                    logging.warning(f"Не удалось распарсить вложение из письма {msg_id}")
                    continue

                # --- Шаг 3: Запись в Google Таблицу ---
                try:
                    # Авторизация в Google Sheets
                    # Создаем словарь с учетными данными из переменной окружения
                    import json
                    service_account_info = json.loads(SERVICE_ACCOUNT_JSON)
                    creds = Credentials.from_service_account_info(
                        service_account_info,
                        scopes=['https://www.googleapis.com/auth/spreadsheets']
                    )
                    client_gs = gspread.authorize(creds)
                    
                    # Открываем таблицу
                    sh = client_gs.open(SPREADSHEET_NAME)
                    ws = sh.worksheet(WORKSHEET_NAME)
                    
                    # Заголовки, если таблица пустая
                    headers = ["RFQ", "Date", "№", "PN", "DESC", "Alt", "R. Qty.", "Unit", "Req.Condition", "Target Date", "Comment"]
                    if not ws.get_all_values():
                        ws.append_row(headers)
                        logging.info("Заголовки добавлены в пустую таблицу")
                    
                    # Добавляем строки
                    for row_dict in parsed_data:
                        row = [row_dict.get(h, "") for h in headers]
                        ws.append_row(row)
                    
                    logging.info(f"Данные из письма {msg_id} успешно записаны в Google Таблицу")
                    
                except Exception as e:
                    logging.error(f"Ошибка при работе с Google Таблицей: {e}", exc_info=True)
                    continue

        logging.info("Скрипт успешно завершил работу")
        
    except Exception as e:
        logging.error(f"Критическая ошибка в основном блоке: {e}", exc_info=True)

if __name__ == "__main__":
    main()
